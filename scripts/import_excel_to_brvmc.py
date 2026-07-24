"""
import_excel_to_brvmc.py — Import Excel → données BRVMC ETF
============================================================
Lit BRVM_Consolidated_Kendall_updated.xlsx et génère :
  - sika_history.json          (prix de clôture historiques)
  - brvm30_index_history.json  (indice BRVMCI — benchmark)
  - sika_societe.json          (nb titres, flottant, capi)
  - dividend_history.json      (dividendes historiques)
  - dashboard_data.json        (w_history BRVMC trimestriel)
  - brvm_composition_latest.json

Usage :
    python import_excel_to_brvmc.py
"""
import os, sys, json
from datetime import datetime, date, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from base import BaseScript

try:
    import openpyxl
except ImportError:
    print("[ERREUR] openpyxl non installé. Lancez : pip install openpyxl")
    sys.exit(1)


EXCEL_PATH = os.path.normpath(os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "..", "BRVM_Consolidated_Kendall_updated.xlsx"
))

# Trimestres BRVM (premier jour ouvré de chaque trimestre depuis 2023)
QUARTER_DATES = [
    "2023-01-02", "2023-04-03", "2023-07-03", "2023-10-02",
    "2024-01-02", "2024-04-02", "2024-07-01", "2024-10-01",
    "2025-01-02", "2025-04-01", "2025-07-01", "2025-10-01",
    "2026-01-02", "2026-04-01", "2026-07-01",
]

START_DATE   = "2023-01-01"
MIN_DAYS_IN  = 20    # nombre minimum de jours de prix pour inclure un ticker au backtest


class ExcelImporter(BaseScript):

    def run(self):
        print(f"Lecture de : {EXCEL_PATH}")
        if not os.path.exists(EXCEL_PATH):
            print(f"[ERREUR] Excel introuvable : {EXCEL_PATH}")
            sys.exit(1)

        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        print("[OK] Excel ouvert.")

        prices   = self._load_prices(wb)
        brvmci   = self._load_brvmci(wb)
        societe  = self._load_societe(wb)
        dividends = self._load_dividends(wb)
        wb.close()

        # Écriture des fichiers
        self._write_sika_history(prices)
        self._write_index_history(brvmci)
        self._write_societe(societe)
        self._write_dividend_history(dividends)
        w_history = self._build_w_history(prices, societe)
        self._write_dashboard_data(w_history, brvmci)
        self._write_composition_latest(prices, societe)

        print("\n[OK] Import terminé. Prêt pour le backtest.")

    # ── Chargement ────────────────────────────────────────────────────────────── #

    def _load_prices(self, wb):
        """Cours_Close → {ticker: {date_str: float}}"""
        ws = wb["📈 Cours_Close"]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        tickers = [h for h in headers[1:] if h]

        prices = {tk: {} for tk in tickers}
        for row in rows[1:]:
            d = row[0]
            if not isinstance(d, datetime):
                continue
            d_str = d.strftime("%Y-%m-%d")
            if d_str < START_DATE:
                continue
            for i, tk in enumerate(tickers):
                val = row[i + 1]
                if val is not None and float(val) > 0:
                    prices[tk][d_str] = float(val)

        # Filtrer les tickers avec trop peu de données
        prices = {tk: v for tk, v in prices.items() if len(v) >= MIN_DAYS_IN}
        print(f"   Cours_Close : {len(prices)} tickers, {sum(len(v) for v in prices.values())} points")
        return prices

    def _load_brvmci(self, wb):
        """BRVM_Indices → {date_str: float} pour .BRVMCI"""
        ws = wb["🏛️ BRVM_Indices"]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        idx_col = headers.index(".BRVMCI") if ".BRVMCI" in headers else None
        if idx_col is None:
            print("[ERREUR] Colonne .BRVMCI introuvable dans BRVM_Indices")
            return {}

        result = {}
        for row in rows[1:]:
            d = row[0]
            if not isinstance(d, datetime):
                continue
            d_str = d.strftime("%Y-%m-%d")
            if d_str < START_DATE:
                continue
            val = row[idx_col]
            if val is not None:
                result[d_str] = float(val)

        print(f"   BRVMCI : {len(result)} dates ({min(result)} a {max(result)})")
        return result

    def _load_societe(self, wb):
        """Capitalisations → {ticker: {nb_titres, flottant_pct, valorisation_mfcfa}}"""
        ws = wb["🏢 Capitalisations"]
        rows = list(ws.iter_rows(values_only=True))
        result = {}
        for row in rows[1:]:
            if not row[0]:
                continue
            tk  = str(row[0]).strip()
            cap_flottante = row[3]   # Cap. Flottante (FCFA)
            cap_globale   = row[4]   # Cap. Globale (FCFA)
            nb_flottants  = row[6]   # Nb Titres Flottants
            nb_total      = row[7]   # Nb Titres Total
            if cap_flottante and nb_flottants and nb_total:
                flottant_pct = float(nb_flottants) / float(nb_total) * 100 if nb_total else 100.0
                result[tk] = {
                    "nb_titres":         float(nb_total),
                    "nb_titres_flottants": float(nb_flottants),
                    "flottant_pct":      round(flottant_pct, 4),
                    "valorisation_mfcfa": round(float(cap_globale) / 1_000_000, 2) if cap_globale else 0,
                    "cap_flottante_fcfa": float(cap_flottante),
                }
        print(f"   Capitalisations : {len(result)} sociétés")
        return result

    def _load_dividends(self, wb):
        """Dividendes_Net → {ticker: {year_str: amount_fcfa}}"""
        ws = wb["💰 Dividendes_Net"]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        years = [str(h) for h in headers[2:] if h is not None]

        result = {}
        for row in rows[1:]:
            if not row[0]:
                continue
            tk = str(row[0]).strip()
            div = {}
            for i, yr in enumerate(years):
                val = row[i + 2]
                if val is not None and float(val) > 0:
                    div[yr] = float(val)
            if div:
                result[tk] = div

        print(f"   Dividendes : {len(result)} tickers")
        return result

    # ── Écriture ─────────────────────────────────────────────────────────────── #

    def _write_sika_history(self, prices):
        out = {}
        for tk, hist in prices.items():
            out[tk] = {d: {"close": v, "volume": 0, "open": v, "high": v, "low": v}
                       for d, v in hist.items()}
        self.save_json("sika_history.json", out)
        print(f"[OK] sika_history.json écrit ({len(out)} tickers)")

    def _write_index_history(self, brvmci):
        self.save_json("brvm30_index_history.json", brvmci)
        print(f"[OK] brvm30_index_history.json écrit ({len(brvmci)} dates)")

    def _write_societe(self, societe):
        # Format attendu par les scripts : nb_titres, flottant_pct, valorisation_mfcfa
        out = {tk: {
            "nb_titres":         v["nb_titres"],
            "flottant_pct":      v["flottant_pct"],
            "valorisation_mfcfa": v["valorisation_mfcfa"],
        } for tk, v in societe.items()}
        self.save_json("sika_societe.json", out)
        print(f"[OK] sika_societe.json écrit ({len(out)} tickers)")

    def _write_dividend_history(self, dividends):
        existing = self.load_json("dividend_history.json", {})
        existing["history"] = dividends
        self.save_json("dividend_history.json", existing)
        print(f"[OK] dividend_history.json écrit ({len(dividends)} tickers)")

    def _build_w_history(self, prices, societe):
        """
        Calcule les poids BRVM Composite à chaque date de rebalancement.
        Poids = cap flottante = nb_titres_flottants × prix_à_la_date
        """
        w_history = {}
        for qd in QUARTER_DATES:
            # Chercher le prix le plus proche de la date (max 5 jours avant)
            cap_map = {}
            for tk, hist in prices.items():
                if tk not in societe:
                    continue
                # Capitalisation totale (comme le BRVMCI officiel)
                nb_total = societe[tk].get("nb_titres", 0)
                if nb_total <= 0:
                    continue
                px = None
                for delta in range(0, 6):
                    d_try = (date.fromisoformat(qd) - timedelta(days=delta)).isoformat()
                    if d_try in hist:
                        px = hist[d_try]
                        break
                if px and px > 0:
                    cap_map[tk] = nb_total * px

            if not cap_map:
                continue

            total_cap = sum(cap_map.values())
            w_history[qd] = {tk: round(cap / total_cap, 6) for tk, cap in cap_map.items()}
            n = len(w_history[qd])
            top3 = sorted(w_history[qd].items(), key=lambda x: -x[1])[:3]
            print(f"   {qd}: {n} titres, top3={[(t, round(w*100,2)) for t,w in top3]}")

        print(f"[OK] w_history calculé ({len(w_history)} trimestres)")
        return w_history

    def _write_dashboard_data(self, w_history, brvmci):
        dd = {
            "etf_name":     "CGF BRVMC ETF",
            "w_history":    w_history,
            "nav_etf":      {},
            "nav_bench":    {},
            "nav_gross":    {},
            "metrics":      {},
            "rebal_history": [],
            "scalability":  {},
            "walk_forward": {},
        }
        self.save_json("dashboard_data.json", dd)
        print(f"[OK] dashboard_data.json écrit")

    def _write_composition_latest(self, prices, societe):
        # Composition BRVMC = tous les tickers actifs (prix dans les 20 derniers jours)
        all_dates = sorted(set(d for hist in prices.values() for d in hist))
        last_date = all_dates[-1] if all_dates else ""
        cutoff = (date.fromisoformat(last_date) - timedelta(days=30)).isoformat() if last_date else ""

        active = []
        for tk, hist in prices.items():
            recent_dates = [d for d in hist if d >= cutoff]
            if recent_dates:
                active.append(tk)
        active.sort()

        comp = {
            "rebal_date":  QUARTER_DATES[-1],
            "scrape_ts":   datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "n_tickers":   len(active),
            "composition": active,
            "entries":     [],
            "exits":       [],
        }
        self.save_json("brvm_composition_latest.json", comp)
        print(f"[OK] brvm_composition_latest.json écrit ({len(active)} tickers actifs)")


if __name__ == "__main__":
    ExcelImporter().run()
